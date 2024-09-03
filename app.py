import openpyxl
import requests
from bs4 import BeautifulSoup
import urllib.parse

# Load the Excel file
file_path = '4BeatsQ1.xlsx'
wb = openpyxl.load_workbook(file_path)
sheet = wb.active

# Define the columns for keywords, longest, and shortest options
keyword_col = 3  # Column C
longest_col = 4  # Column D
shortest_col = 5  # Column E

# Function to search Google and retrieve search result snippets
def get_search_results(keyword):
    query = urllib.parse.quote(keyword)
    url = f"https://www.google.com/search?q={query}"
    
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/93.0.4577.63 Safari/537.36"
    }
    
    response = requests.get(url, headers=headers)
    soup = BeautifulSoup(response.text, 'html.parser')
    
    # Extract all the search result snippets
    snippets = soup.find_all('span', class_='aCOpRe')
    results = [snippet.get_text() for snippet in snippets]
    
    if results:
        longest_result = max(results, key=len)
        shortest_result = min(results, key=len)
        return longest_result, shortest_result
    else:
        return None, None

# Iterate over the keywords in the Excel file and update with results
for row in range(2, sheet.max_row + 1):
    keyword = sheet.cell(row=row, column=keyword_col).value
    
    if keyword:
        print(f"Searching for: {keyword}")
        longest, shortest = get_search_results(keyword)
        
        # Update the Excel sheet with the longest and shortest results
        if longest:
            sheet.cell(row=row, column=longest_col).value = longest
        if shortest:
            sheet.cell(row=row, column=shortest_col).value = shortest

# Save the updated Excel file
wb.save('updated_' + file_path)
print("Excel file updated successfully.")
