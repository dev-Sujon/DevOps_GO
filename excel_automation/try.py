from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
import openpyxl

# Load the Excel workbook
workbook = openpyxl.load_workbook('4BeatsQ1.xlsx')
sheet = workbook.active

# Set up the WebDriver (make sure you have the correct path to your chromedriver)
options = webdriver.ChromeOptions()
options.binary_location = '/usr/bin/chromedriver'
driver = webdriver.Chrome(options=options)

# Function to find the longest and shortest suggestions
def find_longest_and_shortest(suggestions):
    longest = max(suggestions, key=len)
    shortest = min(suggestions, key=len)
    return longest, shortest

for row in range(2, sheet.max_row + 1):
    # Read keyword from Excel
    keyword = sheet.cell(row=row, column=1).value
    
    # Navigate to Google Search
    driver.get('https://www.google.com/')
    
    # Find the search box element and enter the keyword
    search_box = driver.find_element(By.NAME, 'q')
    search_box.send_keys(keyword)
    search_box.send_keys(Keys.RETURN)
    
    time.sleep(2)  # Wait for the search results to load
    
    # Retrieve search suggestions (this part may need to be adjusted based on Google's HTML structure)
    suggestions = driver.find_elements(By.CSS_SELECTOR, 'div.sbl1 span')
    suggestions_text = [suggestion.text for suggestion in suggestions]
    
    if suggestions_text:
        longest, shortest = find_longest_and_shortest(suggestions_text)
        
        # Write the longest and shortest options back into Excel
        sheet.cell(row=row, column=2).value = longest
        sheet.cell(row=row, column=3).value = shortest

# Save the updated Excel file
workbook.save('updated_excel_file.xlsx')

# Close the WebDriver
driver.quit()
